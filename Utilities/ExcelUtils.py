import openpyxl

def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)

def readData(file, sheetName, rowno, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowno, column=columnno).value

def writeData(file, sheetName, rowno, columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rowno, column=columnno).value=data
    workbook.save(file)
    
def getData(file, sheetName, Header,Testcasename):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    row_count = sheet.max_row
    column_count = sheet.max_column
    datamap={}
    for keys in sheet.iter_rows(min_row=1,max_row=1,values_only=True):
        if(keys[0]==Header):
            for i in range(2,row_count+1):
                data_value = sheet.cell(row=i, column=1).value
                if(data_value==Testcasename):
                    for testdata in range(1, column_count+1):
                        cell_obj = sheet.cell(row=i, column=testdata )
#                         print(keys[testdata-1] , cell_obj.value)
                        datamap.update({keys[testdata-1]:cell_obj.value}) 
    return datamap                         
#                     for value in sheet.cell(row=i,max_row=1,values_only=True):
#                         for x in column_count:
#                             datamap.update({keys:value}) 
    #     
#     
#     
#     
#     
#     
#         column_Header_letter = get_column_letter(i)
#         data = sheet.cell(row=i, column=1).value
#         if(data==Testcasename):
#             for j in range(2, column_count+1):
#                 column_letter = get_column_letter(j)
    